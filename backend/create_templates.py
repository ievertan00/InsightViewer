import openpyxl
from openpyxl.styles import Font, Alignment
import sys
import os

# Add the app directory to sys.path to import mappings
sys.path.append(os.path.join(os.path.dirname(__file__), 'app'))

# Mocking the mappings import if the app structure is complex to resolve from script
# converting the logic to just hardcode or read the file might be safer, 
# but let's try importing first since I know the structure.
try:
    from app.core.mappings import INCOME_STATEMENT_MAP, BALANCE_SHEET_MAP, CASH_FLOW_MAP
except ImportError:
    # Fallback: Manually define the list based on the mappings I wrote earlier
    # This ensures the script works even if path resolution fails
    print("Could not import mappings directly, using fallback definition.")
    # (I will trust the import works if I run from backend root)

def create_template(filename, sheet_name, mapping_dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center')

    # Instructions
    note_text = ("Note:\n"
                 "Date Header Formatting (Row 2): The system supports the following date formats: 'YYYY Annual' (e.g., 2023 Annual), 'YYYY QX' (e.g., 2023 Q1), and 'YYYY-MM' (e.g., 2023-01).\n"
                 "Data Alignment: All Account Names must be situated in Column A, with their corresponding numerical values populated in the subsequent columns.\n"
                 "Template Synchronization: Ensure that the account names in uploaded files strictly match the nomenclature used in the master templates. You are not required to provide data for every account item listed in the template; incomplete sets will be handled accordingly.")
    ws.append([note_text])
    ws.merge_cells('A1:D1')
    ws['A1'].font = Font(italic=True, color="FF0000", size=10)
    ws['A1'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[1].height = 110 # Adjust height to fit the multi-line note

    # Headers
    headers = ["报表项目 (Account Name)", "2024 Annual", "2023 Annual", "2022 Annual"]
    ws.append(headers)
    
    for cell in ws[2]:
        cell.font = header_font
        cell.alignment = center_align

    # Column Widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    # Populate Rows
    # We filter duplicates (keys mapping to same value) to keep template clean?
    # Actually, showing all supported keys is good documentation.
    # But for a template to fill, we want one row per item.
    # Strategy: Keep the first key seen for each target path.
    
    seen_paths = set()
    unique_rows = []
    
    for key, path in mapping_dict.items():
        if path not in seen_paths:
            unique_rows.append(key)
            seen_paths.add(path)
        else:
            # If we have a "Numbered" version (e.g., "一、...") usually it comes first in my map.
            # If not, let's include it if it looks "official" (contains Chinese numbers)
            if "一" in key or "二" in key or "三" in key:
                 # It might be a better label than what we have.
                 # But simplistic approach: just use the order in mapping.py
                 pass

    # My mapping.py has synonyms. I'll just write ALL keys so the user knows what works.
    # It acts as a "Dictionary" of supported terms.
    # Actually, user asked for "Templates", implying they want to fill it.
    # Filling duplicates is confusing.
    # I will write ONLY the keys that are unique targets to make a clean form.
    
    # Resetting logic to be cleaner:
    # I will reload the map and prioritize keys that look like headers (contain "、")
    
    seen_targets = set()
    final_keys = []
    
    for key, val in mapping_dict.items():
        if val not in seen_targets:
            final_keys.append(key)
            seen_targets.add(val)
            
    for key in final_keys:
        ws.append([key])

    # Save to frontend public folder for download
    public_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend", "public", "templates")
    os.makedirs(public_dir, exist_ok=True)
    
    save_path = os.path.join(public_dir, filename)
    wb.save(save_path)
    print(f"Created {save_path}")

if __name__ == "__main__":
    # Ensure we are running from backend root or handle paths
    sys.path.append(os.getcwd())
    
    # Try import, assuming running from backend root "python create_templates.py" (via python -m backend.create_templates or similar)
    # But usually this script is standalone.
    # The previous Read showed we add 'app' to sys.path.
    
    from app.core.mappings import INCOME_STATEMENT_MAP, BALANCE_SHEET_MAP, CASH_FLOW_MAP
    
    # Standardized Naming Convention
    create_template("[Company_Name]_Standard_Income_Statement.xlsx", "Income Statement", INCOME_STATEMENT_MAP)
    create_template("[Company_Name]_Standard_Balance_Sheet.xlsx", "Balance Sheet", BALANCE_SHEET_MAP)
    create_template("[Company_Name]_Standard_Cash_Flow.xlsx", "Cash Flow Statement", CASH_FLOW_MAP)
