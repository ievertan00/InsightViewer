import openpyxl
import os

files = [
    "DemoData/康美药业[600518.SH]-利润表.xlsx",
    "DemoData/康美药业[600518.SH]-资产负债表.xlsx"
]

for f_path in files:
    if not os.path.exists(f_path):
        print(f"File not found: {f_path}")
        continue
        
    print(f"\n--- Inspecting {f_path} ---")
    try:
        wb = openpyxl.load_workbook(f_path, data_only=True)
        print(f"Sheet Names: {wb.sheetnames}")
        
        sheet = wb.active
        print(f"Active Sheet: {sheet.title}")
        
        print("First 10 rows:")
        for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
            print(f"Row {i+1}: {row}")
            
    except Exception as e:
        print(f"Error reading file: {e}")
