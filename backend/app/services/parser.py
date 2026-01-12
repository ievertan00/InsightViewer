import re
import io
import openpyxl
from typing import List, Dict, Any, Tuple, Optional
from app.models.schemas import (
    StandardizedReport, Report, FinancialReportData,
    IncomeStatement, BalanceSheet, CashFlowStatement, CompanyMeta
)
from app.core.mappings import INCOME_STATEMENT_MAP, BALANCE_SHEET_MAP, CASH_FLOW_MAP

def detect_sheet_type(sheet_name: str, content_sample: str, filename: str = "") -> Optional[str]:
    """Detects if a sheet is Income, Balance, or Cash Flow."""
    name = sheet_name.lower()
    content = content_sample
    fname = filename.lower()
    
    # 1. Check Sheet Name
    if "利润" in name or "损益" in name or "income" in name:
        return "income_statement"
    if "资产负债" in name or "balance" in name:
        return "balance_sheet"
    if "现金" in name or "cash" in name:
        return "cash_flow_statement"
    
    # 2. Check Content (Relaxed Logic)
    if "营业收入" in content or "营业总收入" in content:
        return "income_statement"
    if "资产" in content and "负债" in content:
        # "资产总计" or "流动资产" usually appear
        return "balance_sheet"
    if "流动资产" in content:
        return "balance_sheet"
    if "经营活动" in content or "现金流量" in content:
        return "cash_flow_statement"
        
    # 3. Fallback: Check Filename (Critical for separated files or encoding issues)
    if "利润" in fname or "income" in fname:
        return "income_statement"
    if "资产负债" in fname or "balance" in fname:
        return "balance_sheet"
    if "现金" in fname or "cash" in fname:
        return "cash_flow_statement"
        
    return None

def find_header_row(sheet, max_rows=30) -> Tuple[int, List[Dict[str, Any]]]:
    """
    Scans the first `max_rows` to find the header row.
    Returns:
        header_row_index (0-based)
        years_map: List of {col_idx: int, year: str}
    """
    header_row_index = -1
    years_map = []
    
    # Heuristic 1: Look for regex dates (2019, 2020-12, 2020.12, 2023年1月)
    # Matches: 2023, 2023-01, 2023.01, 202301, 2023年, 2023年1月
    date_pattern = re.compile(r"20\d{2}(?:[-\./年]\d{1,2}(?:月)?)?|20\d{4}")
    
    for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True)):
        current_row_years = []
        found_date = False
        
        for c_idx, cell_value in enumerate(row):
            str_val = str(cell_value).strip() if cell_value else ""
            # Simple check first
            if "20" in str_val:
                year_match = date_pattern.search(str_val)
                if year_match:
                    year_str = year_match.group(0)
                    # Basic cleanup
                    year_str = year_str.replace("年", "-").replace("月", "")
                    if year_str.endswith("-"): year_str = year_str[:-1]
                    
                    current_row_years.append({"col_idx": c_idx, "year": year_str})
                    found_date = True
        
        if found_date:
            # We found a row with dates. Assume this is the header.
            header_row_index = r_idx
            years_map = current_row_years
            break
            
    # Heuristic 2 (Fallback): Look for "Subject"/"Item" keywords if no dates found
    # (Sometimes columns are "Current Period", "Prior Period" without explicit years in header)
    if header_row_index == -1:
        pass # To be implemented if strict date matching fails
        
    return header_row_index, years_map

def set_nested_value(data_dict: Dict, path: str, value: float):
    """
    Sets a value in a nested dictionary using a dot-notation path.
    e.g., path="total_operating_revenue.operating_revenue"
    """
    keys = path.split('.')
    current = data_dict
    for i, key in enumerate(keys[:-1]):
        if key not in current:
            current[key] = {}
        current = current[key]
        
    last_key = keys[-1]
    current[last_key] = value

def post_process_totals(sheet_type: str, data: Dict):
    """
    Recalculates totals for compound fields if they are 0 but children have values.
    This handles cases where Excel only provides child items but schema expects a total.
    """
    if sheet_type == "balance_sheet":
        # Receivables
        if "current_assets" in data:
            ca = data["current_assets"]
            
            # Notes & Accounts Receivable
            if "notes_and_accounts_receivable" in ca:
                nar = ca["notes_and_accounts_receivable"]
                if nar.get("amount", 0) == 0:
                    nar["amount"] = nar.get("notes_receivable", 0) + nar.get("accounts_receivable", 0)
            
            # Other Receivables
            if "other_receivables_total" in ca:
                ort = ca["other_receivables_total"]
                if ort.get("amount", 0) == 0:
                    ort["amount"] = ort.get("interest_receivable", 0) + ort.get("dividends_receivable", 0) + ort.get("other_receivables", 0)

        # Payables
        if "current_liabilities" in data:
            cl = data["current_liabilities"]
            
            # Notes & Accounts Payable
            if "notes_and_accounts_payable" in cl:
                nap = cl["notes_and_accounts_payable"]
                if nap.get("amount", 0) == 0:
                    nap["amount"] = nap.get("notes_payable", 0) + nap.get("accounts_payable", 0)
            
            # Other Payables
            if "other_payables_total" in cl:
                opt = cl["other_payables_total"]
                if opt.get("amount", 0) == 0:
                    opt["amount"] = opt.get("interest_payable", 0) + opt.get("dividends_payable", 0) + opt.get("other_payables", 0)

def normalize_name(name: str) -> str:
    """
    Normalizes account name by removing:
    1. Whitespace
    2. Serial numbers (e.g., "一、", "1、", "(一)", "（一）")
    3. Accounting prefixes (e.g., "加：", "减：", "其中：")
    4. Suffixes/Keywords like "净额"
    5. Special characters (spaces, :, -, 、)
    
    This ensures "一、加：经营活动产生的现金流量净额 : " matches "经营活动产生的现金流量".
    """
    if not name:
        return ""
        
    # 1. Strip whitespace
    name = name.strip()
    
    # 2. Remove serial numbers
    # Matches (一) or （一）
    name = re.sub(r"^[（\(][一二三四五六七八九十\d]+[）\)]", "", name)
    # Matches 一、 or 1、 or 1.
    name = re.sub(r"^[一二三四五六七八九十\d]+[、\.]", "", name)
    
    # 3. Remove accounting prefixes: 加, 减, 其中 (and optional colon)
    name = re.sub(r"^(加|减|其中)[：:]?", "", name)
    
    # 4. Remove keywords like "净额" and "合计" and content in parentheses
    name = name.replace("净额", "")
    name = name.replace("合计", "")
    name = re.sub(r"[\(（].*?[\)）]", "", name)
    
    # 5. Remove specific punctuation
    name = re.sub(r"[:：\-、]", "", name)
    
    # 6. Remove ALL whitespace (including NBSP, fullwidth, etc.)
    name = "".join(name.split())
    
    return name

import logging

logger = logging.getLogger(__name__)

def parse_sheet_data(sheet, header_row_idx: int, years_map: List[Dict], mapping: Dict) -> Tuple[Dict[str, Dict], List[str]]:
    """
    Parses rows and returns a dict keyed by YEAR containing the structured data.
    Result: ({ "2023": { "total_operating_revenue": { ... } }, ... }, [skipped_rows])
    """
    results_by_year = { item['year']: {} for item in years_map }
    skipped_rows = []
    
    # Pre-calculate normalized mapping keys to improve performance and matching
    # Note: If multiple keys normalize to the same string, the last one wins.
    # Given the specificity of the mapping, this is generally acceptable.
    normalized_mapping = {}
    for k, v in mapping.items():
        norm_k = normalize_name(k)
        if norm_k:
            normalized_mapping[norm_k] = v

    # Iterate rows starting after header
    for row in sheet.iter_rows(min_row=header_row_idx + 2, values_only=True):
        # Assume Column A (index 0) is the Account Name
        raw_account_name = row[0]
        if not raw_account_name:
            continue # Skip completely empty account name rows
            
        account_name = str(raw_account_name).strip()
        
        # Clean account name using the new normalization logic
        clean_name = normalize_name(account_name)
        
        if clean_name in normalized_mapping:
            target_path = normalized_mapping[clean_name]
            
            # Extract value for each year column
            for year_info in years_map:
                col_idx = year_info['col_idx']
                year = year_info['year']
                
                if col_idx < len(row):
                    raw_val = row[col_idx]
                    # Normalize number
                    num_val = 0.0
                    try:
                        if isinstance(raw_val, (int, float)):
                            num_val = float(raw_val)
                        elif isinstance(raw_val, str):
                            # Remove commas, handle parens?
                            raw_val = raw_val.replace(',', '')
                            if raw_val.strip() == '-' or raw_val.strip() == '':
                                num_val = 0.0
                            else:
                                num_val = float(raw_val)
                        elif raw_val is None:
                            num_val = 0.0
                    except ValueError:
                        num_val = 0.0
                        
                    set_nested_value(results_by_year[year], target_path, num_val)
        else:
            # Row name not in mapping, add to exceptions
            if account_name:
                log_msg = f"Mismatch: Raw='{account_name}' -> Normalized='{clean_name}' not found in mapping."
                logger.info(log_msg)
                # Store more detail for the warning report
                skipped_rows.append(f"{account_name} (norm: {clean_name})")
                    
    return results_by_year, skipped_rows

def parse_excel_file(file_content: bytes, filename: str) -> StandardizedReport:
    """
    Main entry point for parsing an Excel file into the StandardizedReport format.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    
    # Initialize container for aggregated year data
    # Structure: { "2023": { "income_statement": {}, "balance_sheet": {}, ... } }
    aggregated_data = {}
    all_warnings = []
    
    # Company Meta Extraction
    # Expected format: "CompanyName_ReportType.xlsx" or "CompanyName-ReportType.xlsx"
    clean_filename = filename.rsplit('.', 1)[0]
    
    company_name = "Unknown"
    if '_' in clean_filename:
        company_name = clean_filename.split('_')[0].strip()
    elif '-' in clean_filename:
        company_name = clean_filename.split('-')[0].strip()
    else:
        company_name = clean_filename.strip()
    
    for sheet_name in wb.sheetnames:
        try:
            sheet = wb[sheet_name]
            
            # 1. Detect Type
            # Sample first few rows for content check
            sample_content = ""
            for r in sheet.iter_rows(max_row=20, values_only=True):
                sample_content += " ".join([str(x) for x in r if x])
                
            sheet_type = detect_sheet_type(sheet_name, sample_content, filename)
            if not sheet_type:
                continue
                
            # 2. Find Header
            header_idx, years_map = find_header_row(sheet)
            if header_idx == -1 or not years_map:
                print(f"Skipping sheet {sheet_name}: No header/years found.")
                continue
                
            # 3. Select Mapping
            mapping = {}
            if sheet_type == "income_statement":
                mapping = INCOME_STATEMENT_MAP
            elif sheet_type == "balance_sheet":
                mapping = BALANCE_SHEET_MAP
            elif sheet_type == "cash_flow_statement":
                mapping = CASH_FLOW_MAP
                
            # 4. Parse Rows
            parsed_years_data, skipped = parse_sheet_data(sheet, header_idx, years_map, mapping)
            
            if skipped:
                all_warnings.append(f"Sheet '{sheet_name}': Skipped {len(skipped)} rows due to no matching mapping: {', '.join(skipped)}")
            
            # 5. Merge into Aggregated Data
            for year, data in parsed_years_data.items():
                # Post-process totals for this year's data
                post_process_totals(sheet_type, data)

                if year not in aggregated_data:
                    aggregated_data[year] = {}
                
                # Init specific report section if not exists
                if sheet_type not in aggregated_data[year]:
                    aggregated_data[year][sheet_type] = {}
                    
                # Merge data (naive merge, assumes structure is built by parse_sheet_data)
                # Since parse_sheet_data returns the full nested dict for that sheet, we can just assign/update
                aggregated_data[year][sheet_type] = data
                
        except Exception as e:
            logger.error(f"Error parsing sheet '{sheet_name}' in file '{filename}': {e}", exc_info=True)
            all_warnings.append(f"Sheet '{sheet_name}': Critical parsing error: {str(e)}")
            continue

    # 6. Construct Final Report Objects
    reports_list = []
    for year, sections in aggregated_data.items():
        
        # Create Pydantic models with default values (0), then update with parsed data
        inc_stmt = IncomeStatement(**sections.get("income_statement", {}))
        bal_sheet = BalanceSheet(**sections.get("balance_sheet", {}))
        cash_flow = CashFlowStatement(**sections.get("cash_flow_statement", {}))
        
        fin_data = FinancialReportData(
            income_statement=inc_stmt,
            balance_sheet=bal_sheet,
            cash_flow_statement=cash_flow
        )
        
        # Determine period type based on year string format
        # e.g., "2023-01", "2023.12" -> Monthly
        # "2023" -> Annual
        p_type = "Annual"
        if "-" in year or "." in year:
            p_type = "Monthly"
        elif len(year) == 6 and year.isdigit(): # 202301
            p_type = "Monthly"
            
        report = Report(
            fiscal_year=year,
            period_type=p_type,
            data=fin_data
        )
        reports_list.append(report)
        
    # Sort reports by year descending
    reports_list.sort(key=lambda x: x.fiscal_year, reverse=True)
    
    return StandardizedReport(
        company_meta=CompanyMeta(name=company_name),
        reports=reports_list,
        parsing_warnings=all_warnings
    )
